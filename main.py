from fastapi import FastAPI, Depends, HTTPException, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import RedirectResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from datetime import datetime
import uvicorn
import openpyxl
import random
import string
from os import remove as delete_files_user
from os.path import isfile

from database import Base, engine, SessionLocal
from models import Employee, Application, PriorityApplication, ApplicationStatus
from schemas import ApplicationCreate, ApplicationOut, StatusUpdate
from utils import hash_password, generate_login
from service import generate_docx, add_employee, archive_employee, update_employee

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

app.mount("/static", StaticFiles(directory="view", html=True), name="static")

Base.metadata.create_all(bind=engine)


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


@app.get("/")
def root():
    return RedirectResponse(url="/static/login.html")


@app.post("/logout")
def logout():
    return RedirectResponse(url="/static/login.html")


@app.post("/auth")
def login(login: str, password: str, db: Session = Depends(get_db)):
    users = db.query(Employee).filter(Employee.login == login).all()
    if not users:
        raise HTTPException(status_code=404, detail="Сотрудник не найден!")
    for user in users:
        if hash_password(password) == user.hashed_password:
            if user.is_archived == 1:
                raise HTTPException(status_code=403, detail="Пользователь в архиве!")
            return {
                "user_id": user.id,
                "full_name": user.full_name,
                "is_admin": user.check_admin
            }
    raise HTTPException(status_code=401, detail="Неверный логин или пароль!")


@app.post("/create_employee")
def create_employee(full_name: str, auto_password: bool, password: str, admin: bool, db: Session = Depends(get_db)):
    if auto_password:
        symbol_pass = string.digits + string.ascii_letters
        password = ''.join(random.choice(symbol_pass) for _ in range(10))
    status = add_employee(db, full_name, password, admin)
    if status == "done":
        return {"message": "Сотрудник добавлен!"}
    else:
        return {"message": "Что-то пошло не так!"}


@app.get("/staff")
def get_staff(db: Session = Depends(get_db)):
    return db.query(Employee).order_by(Employee.is_archived, Employee.full_name).all()


@app.put("/staff/{employee_id}")
def update_staff(employee_id: int, full_name: str, password: str, admin: bool, db: Session = Depends(get_db)):
    employee = db.query(Employee).filter(Employee.id == employee_id).first()
    if not employee:
        return {"message": "Сотрудник не найден."}
    status = update_employee(db, employee_id, full_name, password, admin)
    if status == "done":
        return {"message": "Данные обновлены!"}
    else:
        return {"message": "Что-то пошло не так!"}


@app.patch("/staff/archive/{employee_id}")
def archive_staff(employee_id, db: Session = Depends(get_db)):
    status = archive_employee(db, employee_id)
    if status == "done_archive":
        return {"message": "Сотрудник помещён в архив!"}
    elif status == "done_unarchive":
        return {"message": "Сотрудник востановлен!"}
    elif status == "is_admin":
        return {"message": ("Администратора нельзя архивировать.\n\nЧтобы архивировать, сначала снимите с него статус "
                            "администратора.")}
    else:
        raise HTTPException(status_code=500, detail="Ошибка при архивации")


@app.post("/applications")
def create_application(data: ApplicationCreate, db: Session = Depends(get_db)):
    data_query = Application(
        date_submission=str(datetime.now().strftime("%d.%m.%Y")),
        date_completion="",
        cabinet_number=data.cabinet_number,
        title=data.title,
        problem_description=data.problem_description,
        id_employee=data.id_employee,
        id_priority=data.id_priority,
        id_status=1
    )
    db.add(data_query)
    db.commit()
    db.refresh(data_query)
    return {"message": "Заявка создана", "id": data_query.id}


@app.get("/applications", response_model=list[ApplicationOut])
def get_applications(user_id: int, admin: bool, db: Session = Depends(get_db)):
    if admin:
        query_data_status2 = db.query(Application).where(Application.id_status == 2).order_by(
            desc(Application.id_status),
            desc(func.substr(Application.date_submission, 7, 4) +
                 func.substr(Application.date_submission, 4, 2) +
                 func.substr(Application.date_submission, 1, 2))).all()
        query_data_status13 = db.query(Application).where(Application.id_status != 2).order_by(
            Application.id_status, desc(func.substr(Application.date_submission, 7, 4) +
                                        func.substr(Application.date_submission, 4, 2) +
                                        func.substr(Application.date_submission, 1, 2))).all()
        query_data = query_data_status2 + query_data_status13
    else:
        query_data_status2 = db.query(Application).filter(Application.id_employee == user_id).where(
            Application.id_status == 2).order_by(
            desc(Application.id_status),
            desc(func.substr(Application.date_submission, 7, 4) +
                 func.substr(Application.date_submission, 4, 2) +
                 func.substr(Application.date_submission, 1, 2))).all()
        query_data_status13 = db.query(Application).filter(Application.id_employee == user_id).where(
            Application.id_status != 2).order_by(
            Application.id_status,
            desc(func.substr(Application.date_submission, 7, 4) +
                 func.substr(Application.date_submission, 4, 2) +
                 func.substr(Application.date_submission, 1, 2))).all()
        query_data = query_data_status2 + query_data_status13

    return query_data


@app.get("/applications/{app_id}", response_model=ApplicationOut)
def get_application_by_id(app_id: int, db: Session = Depends(get_db)):
    data_query = db.query(Application).filter(Application.id == app_id).first()
    if not data_query:
        raise HTTPException(status_code=404, detail="Заявка не найдена")
    file_path = f"user_file/generated_application_{data_query.title}.docx"
    if isfile(file_path):
        delete_files_user(file_path)
    return data_query


@app.put("/applications/{app_id}/status")
def update_application_status(app_id: int, data: StatusUpdate, db: Session = Depends(get_db)):
    data_query = db.query(Application).filter(Application.id == app_id).first()
    if not data_query:
        raise HTTPException(status_code=404, detail="Заявка не найдена")
    data_query.id_status = data.new_status
    if data.new_status == 3:
        data_query.date_completion = str(datetime.now().strftime("%d.%m.%Y"))
    db.commit()
    return {"message": "Статус обновлён"}


@app.get("/applications/{id}/export")
def export_docx(id: int, db: Session = Depends(get_db)):
    data_query = db.query(Application).filter(Application.id == id).first()
    name_priority = db.query(PriorityApplication).filter(PriorityApplication.id == int(data_query.id_priority)).first()
    name_status = db.query(ApplicationStatus).filter(ApplicationStatus.id == int(data_query.id_status)).first()
    full_name = db.query(Employee).filter(Employee.id == int(data_query.id_employee)).first()
    output_path = generate_docx(full_name.full_name,
                                name_priority.name,
                                name_status.name,
                                id_application=data_query.id,
                                date_submission=data_query.date_submission,
                                date_completion=data_query.date_completion if data_query.date_completion != "" else "",
                                cabinet_number=data_query.cabinet_number,
                                title=data_query.title,
                                problem_description=data_query.problem_description)

    return FileResponse(output_path,
                        media_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document',
                        filename=f"Заявка_{full_name.full_name}_{data_query.title}_{data_query.date_submission}.docx")


@app.post("/upload_excel")
def upload_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Нужен Excel (.xlsx) файл")

    wb = openpyxl.load_workbook(file.file)
    sheet = wb.active

    if (sheet["A1"].value != "ФИО") or not (sheet["B1"].value is None):
        raise HTTPException(status_code=400, detail="Неверный формат файла!")

    output_wb = openpyxl.Workbook()
    out_ws = output_wb.active
    out_ws.title = "Сотрудники"
    out_ws.append(["ФИО", "Логин", "Пароль"])

    for row in sheet.iter_rows(min_row=2, values_only=True):

        full_name = str(row[0]).strip()
        if not full_name:
            continue

        user_login = generate_login(full_name)

        symbol_pass = string.digits + string.ascii_letters
        password = ''.join(random.choice(symbol_pass) for _ in range(10))

        count = len(db.query(Employee).filter(Employee.login.like(str(f"{user_login}%"))).all())
        if count > 0:
            user_login += str(count)

        employee = Employee(
            full_name=full_name,
            login=user_login,
            hashed_password=hash_password(password),
            check_admin=False
        )
        db.add(employee)
        out_ws.append([full_name, user_login, password])

    db.commit()

    path = "employee_export.xlsx"
    output_wb.save(path)
    return FileResponse(path, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        filename="Сотрудники_с_логинами.xlsx")


if __name__ == "__main__":
    # uvicorn main:app --host 0.0.0.0 --port 8000
    uvicorn.run(
        "main:app",
        host="0.0.0.0",
        port=8000,
        reload=True
    )
